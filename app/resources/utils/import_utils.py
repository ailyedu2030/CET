"""资源导入/导出工具."""

from __future__ import annotations

import json
import logging
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.workbook import Workbook

from app.resources.schemas.resource_schemas import (
    ExamSyllabusBase,
    ImportResult,
    KnowledgePointBase,
    TeachingMaterialBase,
    VocabularyItemBase,
)
from app.shared.models.enums import ContentType, DifficultyLevel

logger = logging.getLogger(__name__)


class ResourceImportUtils:
    """资源导入工具类."""

    @staticmethod
    async def import_from_excel(
        file_path: str, resource_type: str, sheet_name: str | None = None
    ) -> ImportResult:
        """从Excel文件导入资源."""
        try:
            # 读取Excel文件
            df = pd.read_excel(file_path, sheet_name=sheet_name)

            # 根据资源类型选择对应的导入逻辑
            if resource_type == "vocabulary":
                return await ResourceImportUtils._import_vocabulary_from_df(df)
            elif resource_type == "knowledge_point":
                return await ResourceImportUtils._import_knowledge_from_df(df)
            elif resource_type == "material":
                return await ResourceImportUtils._import_material_from_df(df)
            elif resource_type == "syllabus":
                return await ResourceImportUtils._import_syllabus_from_df(df)
            else:
                return ImportResult(
                    total=0,
                    success=0,
                    failed=1,
                    skipped=0,
                    errors=[f"Unsupported resource type: {resource_type}"],
                    warnings=[],
                )

        except Exception as e:
            logger.error(f"Error importing from Excel: {str(e)}")
            return ImportResult(
                total=0,
                success=0,
                failed=1,
                skipped=0,
                errors=[f"Excel import error: {str(e)}"],
                warnings=[],
            )

    @staticmethod
    async def import_from_csv(
        file_path: str, resource_type: str, encoding: str = "utf-8"
    ) -> ImportResult:
        """从CSV文件导入资源."""
        try:
            # 读取CSV文件
            df = pd.read_csv(file_path, encoding=encoding)

            # 根据资源类型选择对应的导入逻辑
            if resource_type == "vocabulary":
                return await ResourceImportUtils._import_vocabulary_from_df(df)
            elif resource_type == "knowledge_point":
                return await ResourceImportUtils._import_knowledge_from_df(df)
            elif resource_type == "material":
                return await ResourceImportUtils._import_material_from_df(df)
            else:
                return ImportResult(
                    total=0,
                    success=0,
                    failed=1,
                    skipped=0,
                    errors=[f"Unsupported resource type: {resource_type}"],
                    warnings=[],
                )

        except Exception as e:
            logger.error(f"Error importing from CSV: {str(e)}")
            return ImportResult(
                total=0,
                success=0,
                failed=1,
                skipped=0,
                errors=[f"CSV import error: {str(e)}"],
                warnings=[],
            )

    @staticmethod
    async def import_from_json(file_path: str, resource_type: str) -> ImportResult:
        """从JSON文件导入资源."""
        try:
            with open(file_path, encoding="utf-8") as f:
                data = json.load(f)

            if not isinstance(data, list):
                data = [data]

            # 根据资源类型选择对应的导入逻辑
            if resource_type == "vocabulary":
                return await ResourceImportUtils._import_vocabulary_from_dict_list(data)
            elif resource_type == "knowledge_point":
                return await ResourceImportUtils._import_knowledge_from_dict_list(data)
            elif resource_type == "material":
                return await ResourceImportUtils._import_material_from_dict_list(data)
            elif resource_type == "syllabus":
                return await ResourceImportUtils._import_syllabus_from_dict_list(data)
            else:
                return ImportResult(
                    total=0,
                    success=0,
                    failed=1,
                    skipped=0,
                    errors=[f"Unsupported resource type: {resource_type}"],
                    warnings=[],
                )

        except Exception as e:
            logger.error(f"Error importing from JSON: {str(e)}")
            return ImportResult(
                total=0,
                success=0,
                failed=1,
                skipped=0,
                errors=[f"JSON import error: {str(e)}"],
                warnings=[],
            )

    @staticmethod
    async def _import_vocabulary_from_df(df: pd.DataFrame) -> ImportResult:
        """从DataFrame导入词汇."""
        result = ImportResult(
            total=len(df),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        required_columns = ["word", "chinese_meaning"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            result.errors.append(f"Missing required columns: {missing_columns}")
            result.failed = result.total
            return result

        for index, row in df.iterrows():
            try:
                # 验证必需字段
                if pd.isna(row["word"]) or pd.isna(row["chinese_meaning"]):
                    result.warnings.append(f"Row {index + 1}: Missing required fields")
                    result.skipped += 1
                    continue

                # 创建词汇对象（这里只是验证，实际创建需要在服务层）
                vocab_data: dict[str, Any] = {
                    "word": str(row["word"]).strip(),
                    "chinese_meaning": str(row["chinese_meaning"]).strip(),
                }

                # 处理可选字段
                optional_fields = [
                    "pronunciation",
                    "part_of_speech",
                    "english_meaning",
                    "learning_tips",
                    "audio_url",
                    "image_url",
                ]
                for field in optional_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        vocab_data[field] = str(row[field]).strip()

                # 处理列表字段
                list_fields = ["synonyms", "antonyms", "tags"]
                for field in list_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        vocab_data[field] = [
                            item.strip() for item in str(row[field]).split(",") if item.strip()
                        ]

                # 处理枚举字段
                if "difficulty_level" in df.columns and not pd.isna(row["difficulty_level"]):
                    try:
                        vocab_data["difficulty_level"] = DifficultyLevel(
                            str(row["difficulty_level"]).lower()
                        )
                    except ValueError:
                        result.warnings.append(
                            f"Row {index + 1}: Invalid difficulty_level, using default"
                        )

                # 处理数值字段
                if "frequency" in df.columns and not pd.isna(row["frequency"]):
                    vocab_data["frequency"] = int(row["frequency"])

                # 处理布尔字段
                if "is_key_word" in df.columns and not pd.isna(row["is_key_word"]):
                    vocab_data["is_key_word"] = bool(row["is_key_word"])

                # 处理复杂字段 - 例句列表
                if "example_sentences" in df.columns and not pd.isna(row["example_sentences"]):
                    # 简单处理：将例句字符串转换为字典列表
                    sentences = str(row["example_sentences"]).split(";")
                    vocab_data["example_sentences"] = [
                        {"sentence": sentence.strip(), "translation": ""}
                        for sentence in sentences
                        if sentence.strip()
                    ]

                # 验证数据（这里只是基础验证）
                VocabularyItemBase(**vocab_data)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Row {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_knowledge_from_df(df: pd.DataFrame) -> ImportResult:
        """从DataFrame导入知识点."""
        result = ImportResult(
            total=len(df),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        required_columns = ["title", "content"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            result.errors.append(f"Missing required columns: {missing_columns}")
            result.failed = result.total
            return result

        for index, row in df.iterrows():
            try:
                # 验证必需字段
                if pd.isna(row["title"]) or pd.isna(row["content"]):
                    result.warnings.append(f"Row {index + 1}: Missing required fields")
                    result.skipped += 1
                    continue

                # 创建知识点对象
                knowledge_data: dict[str, Any] = {
                    "title": str(row["title"]).strip(),
                    "content": str(row["content"]).strip(),
                }

                # 处理可选字段
                optional_fields = ["category", "description"]
                for field in optional_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        knowledge_data[field] = str(row[field]).strip()

                # 处理列表字段
                if "learning_objectives" in df.columns and not pd.isna(row["learning_objectives"]):
                    knowledge_data["learning_objectives"] = [
                        item.strip()
                        for item in str(row["learning_objectives"]).split(",")
                        if item.strip()
                    ]

                if "tags" in df.columns and not pd.isna(row["tags"]):
                    knowledge_data["tags"] = [
                        item.strip() for item in str(row["tags"]).split(",") if item.strip()
                    ]

                # 处理枚举字段
                if "difficulty_level" in df.columns and not pd.isna(row["difficulty_level"]):
                    try:
                        knowledge_data["difficulty_level"] = DifficultyLevel(
                            str(row["difficulty_level"]).lower()
                        )
                    except ValueError:
                        result.warnings.append(
                            f"Row {index + 1}: Invalid difficulty_level, using default"
                        )

                # 处理数值字段
                if "importance_score" in df.columns and not pd.isna(row["importance_score"]):
                    knowledge_data["importance_score"] = float(row["importance_score"])

                if "estimated_time" in df.columns and not pd.isna(row["estimated_time"]):
                    knowledge_data["estimated_time"] = int(row["estimated_time"])

                if "review_frequency" in df.columns and not pd.isna(row["review_frequency"]):
                    knowledge_data["review_frequency"] = int(row["review_frequency"])

                # 处理布尔字段
                if "is_core" in df.columns and not pd.isna(row["is_core"]):
                    knowledge_data["is_core"] = bool(row["is_core"])

                # 处理整数列表字段
                int_list_fields = ["prerequisite_points", "related_points"]
                for field in int_list_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        knowledge_data[field] = [
                            int(item.strip())
                            for item in str(row[field]).split(",")
                            if item.strip().isdigit()
                        ]

                # 处理复杂字段 - 示例、练习、资源
                complex_fields = ["examples", "exercises", "resources"]
                for field in complex_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        # 简单处理：将字符串转换为字典列表
                        items = str(row[field]).split(";")
                        knowledge_data[field] = [
                            {"title": item.strip(), "content": ""} for item in items if item.strip()
                        ]

                # 验证数据
                KnowledgePointBase(**knowledge_data)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Row {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_material_from_df(df: pd.DataFrame) -> ImportResult:
        """从DataFrame导入教材."""
        result = ImportResult(
            total=len(df),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        required_columns = ["title"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            result.errors.append(f"Missing required columns: {missing_columns}")
            result.failed = result.total
            return result

        for index, row in df.iterrows():
            try:
                # 验证必需字段
                if pd.isna(row["title"]):
                    result.warnings.append(f"Row {index + 1}: Missing title")
                    result.skipped += 1
                    continue

                # 创建教材对象
                material_data: dict[str, Any] = {
                    "title": str(row["title"]).strip(),
                }

                # 处理可选字段
                optional_fields = [
                    "publisher",
                    "publication_date",
                    "isbn",
                    "edition",
                    "language",
                    "file_format",
                    "target_audience",
                ]
                for field in optional_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        material_data[field] = str(row[field]).strip()

                # 处理列表字段
                if "authors" in df.columns and not pd.isna(row["authors"]):
                    material_data["authors"] = [
                        item.strip() for item in str(row["authors"]).split(",") if item.strip()
                    ]

                if "learning_objectives" in df.columns and not pd.isna(row["learning_objectives"]):
                    material_data["learning_objectives"] = [
                        item.strip()
                        for item in str(row["learning_objectives"]).split(",")
                        if item.strip()
                    ]

                if "tags" in df.columns and not pd.isna(row["tags"]):
                    material_data["tags"] = [
                        item.strip() for item in str(row["tags"]).split(",") if item.strip()
                    ]

                # 处理枚举字段
                if "content_type" in df.columns and not pd.isna(row["content_type"]):
                    try:
                        material_data["content_type"] = ContentType(
                            str(row["content_type"]).lower()
                        )
                    except ValueError:
                        result.warnings.append(
                            f"Row {index + 1}: Invalid content_type, using default"
                        )

                if "difficulty_level" in df.columns and not pd.isna(row["difficulty_level"]):
                    try:
                        material_data["difficulty_level"] = DifficultyLevel(
                            str(row["difficulty_level"]).lower()
                        )
                    except ValueError:
                        result.warnings.append(
                            f"Row {index + 1}: Invalid difficulty_level, using default"
                        )

                # 处理布尔字段
                bool_fields = ["is_primary", "is_supplementary"]
                for field in bool_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        material_data[field] = bool(row[field])

                # 处理复杂字段 - 章节信息
                if "chapters" in df.columns and not pd.isna(row["chapters"]):
                    # 简单处理：将章节字符串转换为字典列表
                    chapters = str(row["chapters"]).split(";")
                    material_data["chapters"] = [
                        {"title": chapter.strip(), "content": "", "pages": ""}
                        for chapter in chapters
                        if chapter.strip()
                    ]

                # 验证数据
                TeachingMaterialBase(**material_data)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Row {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_syllabus_from_df(df: pd.DataFrame) -> ImportResult:
        """从DataFrame导入考纲."""
        result = ImportResult(
            total=len(df),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        required_columns = ["title", "exam_type"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            result.errors.append(f"Missing required columns: {missing_columns}")
            result.failed = result.total
            return result

        for index, row in df.iterrows():
            try:
                # 验证必需字段
                if pd.isna(row["title"]) or pd.isna(row["exam_type"]):
                    result.warnings.append(f"Row {index + 1}: Missing required fields")
                    result.skipped += 1
                    continue

                # 创建考纲对象
                syllabus_data: dict[str, Any] = {
                    "title": str(row["title"]).strip(),
                    "exam_type": str(row["exam_type"]).strip(),
                }

                # 处理可选字段
                optional_fields = [
                    "exam_level",
                    "version",
                    "effective_date",
                    "expiry_date",
                    "issuing_authority",
                    "description",
                ]
                for field in optional_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        syllabus_data[field] = str(row[field]).strip()

                # 处理列表字段
                if "preparation_suggestions" in df.columns and not pd.isna(
                    row["preparation_suggestions"]
                ):
                    syllabus_data["preparation_suggestions"] = [
                        item.strip()
                        for item in str(row["preparation_suggestions"]).split(",")
                        if item.strip()
                    ]

                if "tags" in df.columns and not pd.isna(row["tags"]):
                    syllabus_data["tags"] = [
                        item.strip() for item in str(row["tags"]).split(",") if item.strip()
                    ]

                # 处理布尔字段
                bool_fields = ["is_current", "is_official"]
                for field in bool_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        syllabus_data[field] = bool(row[field])

                # 处理字典字段
                dict_fields = [
                    "exam_structure",
                    "skill_requirements",
                    "vocabulary_requirements",
                    "scoring_criteria",
                ]
                for field in dict_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        # 简单处理：设置为空字典，实际项目中可能需要JSON解析
                        syllabus_data[field] = {}

                # 处理复杂列表字段
                complex_list_fields = [
                    "grammar_requirements",
                    "topic_areas",
                    "question_types",
                    "sample_papers",
                ]
                for field in complex_list_fields:
                    if field in df.columns and not pd.isna(row[field]):
                        # 简单处理：将字符串转换为字典列表
                        items = str(row[field]).split(";")
                        syllabus_data[field] = [
                            {"title": item.strip(), "description": ""}
                            for item in items
                            if item.strip()
                        ]

                # 验证数据
                ExamSyllabusBase(**syllabus_data)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Row {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_vocabulary_from_dict_list(
        data: list[dict[str, Any]],
    ) -> ImportResult:
        """从字典列表导入词汇."""
        result = ImportResult(
            total=len(data),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        for index, item in enumerate(data):
            try:
                # 验证必需字段
                if "word" not in item or "chinese_meaning" not in item:
                    result.warnings.append(f"Item {index + 1}: Missing required fields")
                    result.skipped += 1
                    continue

                # 验证数据
                VocabularyItemBase(**item)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Item {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_knowledge_from_dict_list(
        data: list[dict[str, Any]],
    ) -> ImportResult:
        """从字典列表导入知识点."""
        result = ImportResult(
            total=len(data),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        for index, item in enumerate(data):
            try:
                # 验证必需字段
                if "title" not in item or "content" not in item:
                    result.warnings.append(f"Item {index + 1}: Missing required fields")
                    result.skipped += 1
                    continue

                # 验证数据
                KnowledgePointBase(**item)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Item {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_material_from_dict_list(
        data: list[dict[str, Any]],
    ) -> ImportResult:
        """从字典列表导入教材."""
        result = ImportResult(
            total=len(data),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        for index, item in enumerate(data):
            try:
                # 验证必需字段
                if "title" not in item:
                    result.warnings.append(f"Item {index + 1}: Missing title")
                    result.skipped += 1
                    continue

                # 验证数据
                TeachingMaterialBase(**item)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Item {index + 1}: {str(e)}")
                result.failed += 1

        return result

    @staticmethod
    async def _import_syllabus_from_dict_list(
        data: list[dict[str, Any]],
    ) -> ImportResult:
        """从字典列表导入考纲."""
        result = ImportResult(
            total=len(data),
            success=0,
            failed=0,
            skipped=0,
            errors=[],
            warnings=[],
        )

        for index, item in enumerate(data):
            try:
                # 验证必需字段
                if "title" not in item or "exam_type" not in item:
                    result.warnings.append(f"Item {index + 1}: Missing required fields")
                    result.skipped += 1
                    continue

                # 验证数据
                ExamSyllabusBase(**item)
                result.success += 1

            except Exception as e:
                result.errors.append(f"Item {index + 1}: {str(e)}")
                result.failed += 1

        return result


class ResourceExportUtils:
    """资源导出工具类."""

    @staticmethod
    async def export_to_excel(
        data: list[dict[str, Any]],
        resource_type: str,
        output_path: str,
        include_metadata: bool = True,
    ) -> str:
        """导出资源到Excel文件."""
        try:
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = resource_type.title()

            if not data:
                workbook.save(output_path)
                return output_path

            # 写入表头
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)

            # 写入数据
            for row, item in enumerate(data, 2):
                for col, header in enumerate(headers, 1):
                    value = item.get(header, "")
                    # 处理列表和字典类型
                    if isinstance(value, list | dict):
                        value = json.dumps(value, ensure_ascii=False)
                    worksheet.cell(row=row, column=col, value=value)

            # 添加元数据工作表
            if include_metadata:
                metadata_sheet = workbook.create_sheet("Metadata")
                metadata_sheet.cell(row=1, column=1, value="Export Time")
                metadata_sheet.cell(row=1, column=2, value=str(pd.Timestamp.now()))
                metadata_sheet.cell(row=2, column=1, value="Resource Type")
                metadata_sheet.cell(row=2, column=2, value=resource_type)
                metadata_sheet.cell(row=3, column=1, value="Total Records")
                metadata_sheet.cell(row=3, column=2, value=len(data))

            workbook.save(output_path)
            return output_path

        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise

    @staticmethod
    async def export_to_csv(
        data: list[dict[str, Any]], output_path: str, encoding: str = "utf-8"
    ) -> str:
        """导出资源到CSV文件."""
        try:
            if not data:
                Path(output_path).touch()
                return output_path

            df = pd.DataFrame(data)

            # 处理列表和字典类型的数据
            for column in df.columns:
                df[column] = df[column].apply(
                    lambda x: (
                        json.dumps(x, ensure_ascii=False) if isinstance(x, list | dict) else x
                    )
                )

            df.to_csv(output_path, index=False, encoding=encoding)
            return output_path

        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            raise

    @staticmethod
    async def export_to_json(
        data: list[dict[str, Any]], output_path: str, include_metadata: bool = True
    ) -> str:
        """导出资源到JSON文件."""
        try:
            export_data = {
                "data": data,
            }

            if include_metadata:
                export_data["metadata"] = {
                    "export_time": str(pd.Timestamp.now()),
                    "total_records": len(data),
                }

            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)

            return output_path

        except Exception as e:
            logger.error(f"Error exporting to JSON: {str(e)}")
            raise
